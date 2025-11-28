# utils.py
from io import BytesIO
from decimal import Decimal
from datetime import datetime

from openpyxl import load_workbook
from django.db import transaction

from .models import Calificacion


def importar_calificaciones_desde_excel(uploaded_file):
    """
    Lee un archivo Excel (.xlsx) subido desde Django
    y crea/actualiza registros Calificacion.
    """

    # Cargar el Excel en memoria (no se guarda en disco)
    wb = load_workbook(filename=BytesIO(uploaded_file.read()), data_only=True)
    ws = wb.active  # primera hoja

    # Leer encabezados de la fila 1 (A1, B1, C1, ...)
    headers = [cell.value for cell in ws[1]]

    # Ejemplo: headers = ['ejercicio','mercado','instrumento','fecha',...,'factor_37']

    # Transaction.atomic → si algo falla, no se guarda nada
    with transaction.atomic():
        # Recorremos desde la fila 2 hasta la última
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Si la fila está completamente vacía, la saltamos
            if all(value is None for value in row):
                continue

            # Convertir la fila en dict: {'ejercicio': 2023, 'mercado': 'ACN', ...}
            fila = dict(zip(headers, row))

            try:
                calificacion_data = limpiar_y_convertir_fila(fila)
            except Exception as e:
                # Si hay un error en esa fila, lanzamos excepción con detalle
                raise ValueError(f"Error en fila con datos {fila}: {e}")

            # Aquí decides si crear siempre o hacer update_or_create.
            # Ejemplo: actualizar si ya existe combinación (ejercicio, mercado, instrumento, secuencia)

            Calificacion.objects.update_or_create(
                ejercicio=calificacion_data["ejercicio"],
                mercado=calificacion_data["mercado"],
                instrumento=calificacion_data["instrumento"],
                secuencia=calificacion_data["secuencia"],
                defaults=calificacion_data,
            )

def limpiar_y_convertir_fila(fila: dict) -> dict:
    """
    Recibe un dict tal como viene del Excel y lo transforma al formato
    que necesita el modelo Calificacion (tipos correctos).
    """

    # --- Campos básicos ---
    ejercicio = int(fila["ejercicio"])
    mercado = str(fila["mercado"]).strip()
    instrumento = str(fila["instrumento"]).strip()

    # fecha puede venir como datetime (tipo fecha) o como string
    fecha_valor = fila["fecha"]
    if isinstance(fecha_valor, datetime):
        fecha = fecha_valor.date()
    else:
        # asumimos string 'dd-mm-aaaa'
        fecha = datetime.strptime(str(fecha_valor), "%d-%m-%Y").date()

    secuencia = int(fila["secuencia"])
    numero_dividendo = int(fila["numero_dividendo"])
    tipo_sociedad = str(fila["tipo_sociedad"]).strip()

    # valor_historico puede venir como float o Decimal
    valor_historico = Decimal(str(fila["valor_historico"]))

    data = {
        "ejercicio": ejercicio,
        "mercado": mercado,
        "instrumento": instrumento,
        "fecha": fecha,
        "secuencia": secuencia,
        "numero_dividendo": numero_dividendo,
        "tipo_sociedad": tipo_sociedad,
        "valor_historico": valor_historico,
    }

    # --- Factores: factor_8 ... factor_37 ---
    for n in range(8, 38):  # del 8 al 37
        nombre_columna = f"factor_{n}"

        if nombre_columna not in fila:
            # si no existe la columna en el Excel, lo dejamos en None
            data[nombre_columna] = None
            continue

        valor = fila[nombre_columna]

        if valor is None:
            data[nombre_columna] = None
        else:
            data[nombre_columna] = Decimal(str(valor))

    return data